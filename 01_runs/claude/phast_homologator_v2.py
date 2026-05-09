"""
phast_homologator_v2.py
=======================
PHAST Pseudocomponent Homologator — phast_pseudomps  [v2]

Cambios respecto a v1
---------------------
★ WATER OVERRIDE (nuevo)
    Antes de homologar, se escanean todas las corrientes en busca de agua
    (H2O / Water / Agua) con fracción molar o másica superior a un umbral
    configurable (default 0.80 = 80%).

    Las corrientes detectadas se muestran en terminal con %mol y %w del agua,
    y el usuario decide a cuáles aplicar el override:

        • El agua se separa del stream antes de homologar.
        • El resto de componentes se homologa y renormaliza dentro del espacio
          (1 - x_agua_original).
        • El agua se reintegra AL FINAL con su fracción original intacta.
        • Resultado garantizado: x_agua + Σ_resto = 1.000000 (sin inflación).

    Umbral configurable:
        --water-threshold 0.80   (línea de comando)
        o se solicita interactivamente si se omite.

Usage:
    python phast_homologator_v2.py                        # modo interactivo
    python phast_homologator_v2.py --n 14                 # N fijo, resto interactivo
    python phast_homologator_v2.py --input f.xlsx \\
        --output result.xlsx --n 12 --water-threshold 0.85

Resolution priority (per component):
    1. Pure-compound alias match  → canonical PHAST name, short-circuits MW search.
    2. Pseudocomponent detection  → restricts pool to hydrocarbon-only (CnHm).
    3. MW nearest-neighbour       → within selected pool; tie-break via alias.

Key guarantees:
    - Output ALWAYS sums to 100 %mol and 100 %w (full renormalisation).
    - Water-override streams: agua fracción = original; resto renorm en (1-x_agua).
    - N is requested interactively or via --n flag (1-18, default 14).
"""

from __future__ import annotations

import argparse, json, math, re, sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────
LB_PER_KG               = 2.20462262
KG_PER_LB               = 0.45359237
AUTO_THRESHOLD           = 1e-5          # 0.001 % absolute floor
PHAST_MAX                = 18            # PHAST hard limit per stream
MW_WARN_PCT              = 5.0           # warn when MW match error exceeds this %
WATER_OVERRIDE_DEFAULT   = 0.80         # default threshold (can be overridden by user)

# Water name variants recognised across Spanish/English/formula
WATER_KEYS = frozenset({'water', 'h2o', 'agua'})

# ── Pseudocomponent patterns ───────────────────────────────────────────────────
PSEUDOCOMPONENT_PATTERNS = (
    re.compile(r"^\d+\s*-\s*\d+\s*c\*$",  re.IGNORECASE),
    re.compile(r"^\d+\+\s*c\*$",           re.IGNORECASE),
    re.compile(r"^nbp.*$",                  re.IGNORECASE),
    re.compile(r"^c(?:\d+|n)\+$",          re.IGNORECASE),
)

_HYDROCARBON_FORMULA_RE = re.compile(r"^C(?:\d+)?H\d+$", re.IGNORECASE)


def _is_pseudocomponent(name: str) -> bool:
    stripped = name.strip()
    return any(p.match(stripped) for p in PSEUDOCOMPONENT_PATTERNS)


def _is_hydrocarbon(formula: str | None) -> bool:
    if not formula:
        return False
    return bool(_HYDROCARBON_FORMULA_RE.fullmatch(formula.replace(" ", "")))


def _is_water(name: str) -> bool:
    """Returns True if the component name refers to water."""
    return name.strip().lower() in WATER_KEYS


# ── Built-in HYSYS / simulator alias table ────────────────────────────────────
_HYSYS: dict[str, str] = {
    "C1": "Methane",  "C2": "Ethane",   "C3": "Propane",
    "C4": "n-Butane", "C5": "n-Pentane","C6": "n-Hexane",
    "C7": "n-Heptane","C8": "n-Octane", "C9": "n-Nonane","C10": "n-Decane",
    "Ar": "Argon", "AR": "Argon",
    "H2O": "Water",  "H2S": "Hydrogen sulfide",
    "CO":  "Carbon monoxide", "CO2": "Carbon dioxide",
    "NH3": "Ammonia",         "HCN": "Hydrogen cyanide",
    "N2":  "Nitrogen",        "O2":  "Oxygen",
    "i-C4": "Isobutane",  "n-C4": "n-Butane",
    "i-C5": "Isopentane", "n-C5": "n-Pentane",
    "22-Mpropane": "Neopentane", "neo-pentane": "Neopentane",
    "Methane":"Methane","Ethane":"Ethane","Propane":"Propane",
    "Isobutane":"Isobutane","Isopentane":"Isopentane",
    "n-Butane":"n-Butane","n-Pentane":"n-Pentane","n-Hexane":"n-Hexane",
    "n-Heptane":"n-Heptane","n-Octane":"n-Octane","n-Nonane":"n-Nonane",
    "n-Decane":"n-Decane","Hydrogen":"Hydrogen","Nitrogen":"Nitrogen",
    "Oxygen":"Oxygen","Argon":"Argon","Water":"Water",
    "Metano":"Methane","Etano":"Ethane","Propano":"Propane",
    "Nitrógeno":"Nitrogen","Nitrogeno":"Nitrogen",
    "Oxígeno":"Oxygen","Oxigeno":"Oxygen","Agua":"Water",
}


# ── Data model ─────────────────────────────────────────────────────────────────

@dataclass
class PhastCompound:
    phast_name:           str
    mw:                   float
    formula:              str | None = None
    cas:                  str | None = None
    status:               int = 1
    flammable_toxic_flag: int | None = None
    is_toxic:             bool = False
    is_flammable:         bool = False
    is_inert:             bool = False
    idlh_ppm:             float | None = None
    erpg_2_ppm:           float | None = None
    stel_ppm:             float | None = None
    lfl_ppm:              float | None = None
    ufl_ppm:              float | None = None
    heat_of_comb_kJ_kmol: float | None = None


@dataclass
class RawComponent:
    name:        str
    mass_kgh:    float
    molar_kmolh: float

    @property
    def mw_calc(self) -> float:
        """
        Calculated molecular weight (g/mol) from mass and molar flows.

        Returns 0.0 when molar_kmolh is zero (or negative) so that the
        downstream resolver can route to the 'fallback_no_mw' path instead
        of raising ZeroDivisionError.  A value of 0.0 is treated as "MW
        unknown" throughout resolve() and homologate().
        """
        if self.molar_kmolh <= 0:
            return 0.0
        return self.mass_kgh / self.molar_kmolh


@dataclass
class StreamInput:
    project:    str
    escenario:  object
    corriente:  object
    mass_unit:  str
    mol_unit:   str
    components: list[RawComponent] = field(default_factory=list)


@dataclass
class HomologatedRow:
    source_name:        str
    phast_name:         str
    mw_calc:            float
    phast_mw:           float
    mass_kgh:           float
    molar_kmolh:        float
    molar_frac:         float
    is_pseudocomponent: bool
    match_pool:         str
    match_method:       str
    mw_error_pct:       float
    warning:            str


@dataclass
class ConsolidatedRow:
    rank:               int
    phast_name:         str
    phast_mw:           float
    phast_status:       int
    molar_frac_raw:     float
    molar_frac_pct_raw: float
    frac_corrected:     float
    frac_corrected_pct: float
    mass_kgh:           float
    mass_lbh:           float
    mass_pct_w:         float


@dataclass
class StreamResult:
    stream:              StreamInput
    homologated:         list[HomologatedRow]
    consolidated:        list[ConsolidatedRow]
    total_mass_kgh:      float
    total_molar_kmolh:   float
    n_requested:         int
    n_used:              int
    threshold:           float
    frac_retained:       float
    molar_sum_check:     float
    water_override:      bool  = False   # ★ NEW: flag indicating override was applied
    water_frac_original: float = 0.0    # ★ NEW: original molar fraction of water


# ══════════════════════════════════════════════════════════════════════════════
# ★ NEW — WATER OVERRIDE PRE-SCAN
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class WaterScanResult:
    """Holds water content metrics for a single stream."""
    stream_idx:   int
    escenario:    object
    corriente:    object
    water_mol_frac:  float   # molar fraction of water
    water_mass_frac: float   # mass  fraction of water
    water_kmolh:     float
    water_kgh:       float
    total_kmolh:     float
    total_kgh:       float


def scan_water_content(streams: list[StreamInput]) -> list[WaterScanResult]:
    """
    Scan all streams and return WaterScanResult for EVERY stream
    (regardless of threshold — the caller filters by threshold).
    """
    results: list[WaterScanResult] = []
    for idx, s in enumerate(streams):
        total_kmolh = sum(c.molar_kmolh for c in s.components)
        total_kgh   = sum(c.mass_kgh    for c in s.components)
        w_kmolh = sum(c.molar_kmolh for c in s.components if _is_water(c.name))
        w_kgh   = sum(c.mass_kgh    for c in s.components if _is_water(c.name))
        mol_frac  = w_kmolh / total_kmolh if total_kmolh > 0 else 0.0
        mass_frac = w_kgh   / total_kgh   if total_kgh   > 0 else 0.0
        results.append(WaterScanResult(
            stream_idx=idx, escenario=s.escenario, corriente=s.corriente,
            water_mol_frac=mol_frac, water_mass_frac=mass_frac,
            water_kmolh=w_kmolh, water_kgh=w_kgh,
            total_kmolh=total_kmolh, total_kgh=total_kgh,
        ))
    return results


def prompt_water_override(
    scan_results: list[WaterScanResult],
    threshold:    float,
) -> set[int]:
    """
    Display streams that exceed the water threshold and let the user decide
    which ones to apply the Water Override to.

    Returns a set of stream indices that will receive the override.
    """
    flagged = [r for r in scan_results if r.water_mol_frac >= threshold]

    if not flagged:
        print(f"\n✓ Ninguna corriente supera el umbral de agua ({threshold*100:.0f}% molar).")
        return set()

    # ── Print detection table ──────────────────────────────────────────────
    print()
    print("═" * 74)
    print(f"  WATER OVERRIDE — Corrientes con contenido de agua ≥ {threshold*100:.1f}%")
    print("═" * 74)
    print(f"  {'No.':<4}  {'Escenario':<12}  {'Corriente':<20}  "
          f"{'%Mol agua':>10}  {'%Másico agua':>13}  {'Agua (kgmol/h)':>15}  {'Agua (kg/h)':>12}")
    print("  " + "─" * 70)
    for pos, r in enumerate(flagged, 1):
        print(f"  [{pos:<2}]  {str(r.escenario):<12}  {str(r.corriente):<20}  "
              f"{r.water_mol_frac*100:>9.2f}%  {r.water_mass_frac*100:>12.2f}%  "
              f"{r.water_kmolh:>15.4f}  {r.water_kgh:>12.2f}")
    print()

    print("  WATER OVERRIDE: el agua pasará con su fracción original intacta.")
    print("  El resto de componentes se renormalizará en el espacio (1 - x_agua).")
    print()
    raw = input(
        "  ¿A cuáles corrientes aplicar el override?\n"
        "  (números, ej: 1,3  |  Enter = TODAS  |  0 = ninguna): "
    ).strip()

    if raw == "0":
        print("  Sin override aplicado.")
        return set()

    if raw == "" or raw.lower() in ("todas", "all"):
        selected = flagged
    else:
        try:
            indices = [int(x.strip()) - 1 for x in raw.split(",")]
            selected = [flagged[i] for i in indices if 0 <= i < len(flagged)]
        except Exception:
            print("  Entrada inválida — se aplicará override a TODAS las detectadas.")
            selected = flagged

    if selected:
        names = [f"Esc{r.escenario}/C{r.corriente}" for r in selected]
        print(f"\n  ✓ Override habilitado para: {', '.join(names)}")
    else:
        print("  Sin override aplicado.")

    return {r.stream_idx for r in selected}


# ── DB & alias loading ─────────────────────────────────────────────────────────

def load_db(path: str | Path) -> tuple[list[PhastCompound], list[PhastCompound]]:
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)

    all_compounds: list[PhastCompound] = []
    for item in raw:
        props = item.get('properties', {})
        flag  = props.get('flammable_toxic_flag') if props else item.get('flammable_toxic_flag')
        if flag is None:
            flag = item.get('flammable_toxic_flag')
        try: flag = int(flag) if flag is not None else None
        except: flag = None
        all_compounds.append(PhastCompound(
            phast_name           = item['phast_name'],
            mw                   = float(item['mw']),
            formula              = item.get('formula'),
            cas                  = item.get('cas'),
            status               = item.get('status', 1),
            flammable_toxic_flag = flag,
            is_toxic             = item.get('is_toxic', flag in (-1, 0) if flag is not None else False),
            is_flammable         = item.get('is_flammable', flag in (0, 1) if flag is not None else False),
            is_inert             = item.get('is_inert', flag == -2 if flag is not None else False),
            idlh_ppm             = props.get('idlh_ppm')             if props else None,
            erpg_2_ppm           = props.get('erpg_2_ppm')           if props else None,
            stel_ppm             = props.get('stel_ppm')             if props else None,
            lfl_ppm              = props.get('lower_flammability_limit_ppm') if props else None,
            ufl_ppm              = props.get('upper_flammability_limit_ppm') if props else None,
            heat_of_comb_kJ_kmol = props.get('heat_of_combustion_kJ_kmol')  if props else None,
        ))

    all_compounds.sort(key=lambda c: (c.mw, c.phast_name))
    hydrocarbon_only = [c for c in all_compounds if _is_hydrocarbon(c.formula)]
    return all_compounds, hydrocarbon_only


def load_aliases(path: str | Path) -> dict[str, str]:
    with open(path, encoding='utf-8') as f:
        ext = json.load(f)
    merged = dict(_HYSYS)
    merged.update(ext)

    _TARGET_REDIRECTS: dict[str, str] = {
        "1-Pentene": "n-Pentane",
    }
    for key, target in list(merged.items()):
        if target in _TARGET_REDIRECTS:
            merged[key] = _TARGET_REDIRECTS[target]

    return merged


# ── MW nearest-neighbour ───────────────────────────────────────────────────────

def _find_closest(mw: float, pool: list[PhastCompound]) -> list[PhastCompound]:
    if not pool:
        return []
    diffs = [(c, abs(mw - c.mw)) for c in pool]
    min_diff = min(d for _, d in diffs)
    tol = 1e-12
    return [c for c, d in diffs if abs(d - min_diff) <= tol]


def _resolve_tie(
    candidates:      list[PhastCompound],
    alias_candidate: str | None,
) -> tuple[PhastCompound, str]:
    if alias_candidate:
        for c in candidates:
            if c.phast_name == alias_candidate:
                return c, "mw_tie_resolved_by_alias"
    return sorted(candidates, key=lambda c: (c.mw, c.phast_name))[0], "mw_tie_unresolved"


# ── Component resolver ─────────────────────────────────────────────────────────

def resolve(
    name:             str,
    mw_calc:          float,
    all_compounds:    list[PhastCompound],
    hydrocarbon_only: list[PhastCompound],
    aliases:          dict[str, str],
) -> dict:
    alias_candidate: str | None = aliases.get(name) or aliases.get(name.strip())
    if not alias_candidate:
        nl = name.lower().strip()
        for k, v in aliases.items():
            if k.lower() == nl:
                alias_candidate = v
                break

    if alias_candidate:
        match = next((c for c in all_compounds if c.phast_name == alias_candidate), None)
        if match:
            mw_err = abs(mw_calc - match.mw) / mw_calc * 100 if mw_calc > 0 else 0.0
            return {
                "phast_name":        match.phast_name,
                "phast_mw":          match.mw,
                "phast_status":      match.status,
                "is_pseudocomponent":False,
                "match_pool":        "full_catalog",
                "match_method":      "alias",
                "mw_error_pct":      round(mw_err, 4),
                "warning":           (f"MW match error {mw_err:.2f}% via alias."
                                      if mw_err > MW_WARN_PCT else ""),
                "alias_candidate":   alias_candidate,
            }

    is_pseudo = _is_pseudocomponent(name)
    pool      = hydrocarbon_only if is_pseudo else all_compounds
    pool_name = "hydrocarbon_only" if is_pseudo else "full_catalog"

    if mw_calc <= 0 or math.isnan(mw_calc):
        match   = pool[0] if pool else all_compounds[0]
        method  = "fallback_no_mw"
        if math.isnan(mw_calc):
            warning = (
                "MW is NaN (check input flows); "
                "assigned to lightest compound in pool."
            )
        else:
            # mw_calc <= 0: molar flow was zero or missing
            warning = (
                "Molar flow is zero or negative; MW cannot be calculated. "
                "Assigned to lightest compound in pool."
            )
    else:
        candidates = _find_closest(mw_calc, pool)
        if len(candidates) == 1:
            match, method = candidates[0], "mw_unique"
        else:
            match, method = _resolve_tie(candidates, alias_candidate)

        mw_err  = abs(mw_calc - match.mw) / mw_calc * 100 if mw_calc > 0 else 0.0
        warning = f"MW match error {mw_err:.2f}%." if mw_err > MW_WARN_PCT else ""

    mw_err = abs(mw_calc - match.mw) / mw_calc * 100 if mw_calc > 0 else 0.0

    return {
        "phast_name":        match.phast_name,
        "phast_mw":          match.mw,
        "phast_status":      match.status,
        "is_pseudocomponent":is_pseudo,
        "match_pool":        pool_name,
        "match_method":      method,
        "mw_error_pct":      round(mw_err, 4),
        "warning":           warning,
        "alias_candidate":   alias_candidate,
    }


# ── Unit conversion ────────────────────────────────────────────────────────────

def to_kgh(v: float, unit: str) -> float:
    u = unit.lower().replace(' ', '').replace('/', '')
    if u in ('kgh', 'kg/h'):   return v
    if u in ('gh',  'g/h'):    return v / 1000.0
    if u in ('lbh', 'lb/h'):   return v * KG_PER_LB
    return v

def to_kmolh(v: float, unit: str) -> float:
    u = unit.lower().replace(' ', '').replace('/', '')
    if u in ('kgmolh', 'kmolh', 'kgmol/h', 'kmol/h'): return v
    if u in ('molh',   'mol/h'):                         return v / 1000.0
    if u in ('lbmolh', 'lbmol/h'):                      return v * KG_PER_LB
    return v


# ── Input parser ───────────────────────────────────────────────────────────────

def parse_input(path: str | Path) -> list[StreamInput]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['input_template']

    scols: list[tuple[int, int]] = []
    col = 2
    while col <= ws.max_column:
        h = ws.cell(1, col).value
        if h and 'corriente' in str(h).lower():
            scols.append((col, col + 1)); col += 2
        else:
            col += 1

    comp_rows: list[tuple[int, str]] = []
    for row in range(9, ws.max_row + 1):
        n = ws.cell(row, 1).value
        if n and str(n).strip():
            comp_rows.append((row, str(n).strip()))

    def clean(v) -> float:
        if v is None: return 0.0
        if isinstance(v, str):
            s = v.strip().lower()
            if s in ('no flow', 'n/a', 'na', '-', ''): return 0.0
            try:    return float(s)
            except: return 0.0
        try:    return float(v)
        except: return 0.0

    streams: list[StreamInput] = []
    for mc, mlc in scols:
        esc  = ws.cell(3, mc).value
        corr = ws.cell(4, mc).value
        if esc is None and corr is None:
            continue
        mu   = str(ws.cell(5, mc).value or 'kg/h')
        molu = str(ws.cell(6, mc).value or 'kgmol/h')
        proj = str(ws.cell(2, mc).value or 'N/A')

        comps: list[RawComponent] = []
        for row, name in comp_rows:
            m  = to_kgh(   clean(ws.cell(row, mc ).value), mu  )
            mo = to_kmolh( clean(ws.cell(row, mlc).value), molu)
            if m > 0 or mo > 0:
                comps.append(RawComponent(name=name, mass_kgh=m, molar_kmolh=mo))

        if comps:
            streams.append(StreamInput(
                project=proj, escenario=esc, corriente=corr,
                mass_unit=mu, mol_unit=molu, components=comps,
            ))

    return streams


# ── Homologation engine ────────────────────────────────────────────────────────

def homologate(
    stream:              StreamInput,
    all_compounds:       list[PhastCompound],
    hydrocarbon_only:    list[PhastCompound],
    aliases:             dict[str, str],
    n_requested:         int,
    water_override:      bool  = False,   # ★ NEW
    water_frac_original: float = 0.0,    # ★ NEW — molar fraction of water in ORIGINAL stream
) -> StreamResult:
    """
    Full homologation pipeline.

    Water Override mode (water_override=True):
        - 'stream' received here is ALREADY the dry sub-stream (water removed by caller).
        - After renormalising the dry sub-stream to Σ=1, each fraction is scaled by
          (1 - water_frac_original) so that the dry components occupy only their
          correct share of the total.
        - A synthetic Water ConsolidatedRow is appended at the end with
          frac_corrected = water_frac_original.
        - Final check: Σ frac_corrected = 1.000000 ✓
    """
    total_kmolh = sum(c.molar_kmolh for c in stream.components)
    total_kgh   = sum(c.mass_kgh    for c in stream.components)

    if total_kmolh <= 0:
        return StreamResult(stream, [], [], total_kgh, 0,
                            n_requested, 0, 0.0, 0.0, 0.0,
                            water_override, water_frac_original)

    # ── Resolve each component ────────────────────────────────────────────
    rows: list[HomologatedRow] = []
    for c in stream.components:
        r = resolve(c.name, c.mw_calc, all_compounds, hydrocarbon_only, aliases)
        rows.append(HomologatedRow(
            source_name        = c.name,
            phast_name         = r['phast_name'],
            mw_calc            = round(c.mw_calc, 4),
            phast_mw           = r['phast_mw'],
            mass_kgh           = round(c.mass_kgh,    6),
            molar_kmolh        = round(c.molar_kmolh, 6),
            molar_frac         = round(c.molar_kmolh / total_kmolh, 8),
            is_pseudocomponent = r['is_pseudocomponent'],
            match_pool         = r['match_pool'],
            match_method       = r['match_method'],
            mw_error_pct       = r['mw_error_pct'],
            warning            = r['warning'],
        ))

    # ── Consolidate by PHAST compound ─────────────────────────────────────
    groups: dict[str, dict] = {}
    for r in rows:
        if r.phast_name not in groups:
            ph = next((c for c in all_compounds if c.phast_name == r.phast_name), None)
            groups[r.phast_name] = {
                'phast_name':   r.phast_name,
                'phast_mw':     ph.mw     if ph else r.phast_mw,
                'phast_status': ph.status if ph else 1,
                'molar_kmolh':  0.0,
            }
        groups[r.phast_name]['molar_kmolh'] += r.molar_kmolh

    # ── Rank descending ───────────────────────────────────────────────────
    ranked = sorted(groups.values(), key=lambda g: g['molar_kmolh'], reverse=True)
    active = [g for g in ranked if g['molar_kmolh'] > 0]

    # ── Threshold = frac of Nth compound ─────────────────────────────────
    # In water override mode, Water occupies one of the requested PHAST slots
    # even though it is appended after dry-component renormalisation.
    water_slot = 1 if water_override and water_frac_original > 0 else 0
    dry_requested = max(0, n_requested - water_slot)
    N = min(dry_requested, len(active), PHAST_MAX - water_slot)
    threshold = (
        max(active[N - 1]['molar_kmolh'] / total_kmolh, AUTO_THRESHOLD)
        if N > 0 else math.inf
    )

    # ── Filter ────────────────────────────────────────────────────────────
    included = [
        (g, g['molar_kmolh'] / total_kmolh)
        for g in ranked
        if g['molar_kmolh'] / total_kmolh >= threshold
    ]
    frac_retained = sum(f for _, f in included)

    # ── Warn when a toxic compound is excluded by the N-slot threshold ────
    # This surfaces the hazard in Part A of the Excel output so it is
    # visible even when the user skips the interactive post-options step.
    included_names = {g['phast_name'] for g, _ in included}
    _compound_lookup = {c.phast_name: c for c in all_compounds}
    for hr in rows:
        if hr.phast_name in included_names:
            continue  # compound survived → no warning needed
        phast_c = _compound_lookup.get(hr.phast_name)
        if phast_c is not None and phast_c.is_toxic:
            toxic_note = (
                f"HAZARD: toxic compound excluded by N={n_requested} slot limit "
                f"(molar fraction {hr.molar_frac:.6f} < threshold {threshold:.6f}). "
                "Consider post-homologation options."
            )
            # Append to any existing warning rather than overwriting it.
            hr.warning = (hr.warning + "  " + toxic_note).strip() if hr.warning else toxic_note

    # ── Renormalise dry components ────────────────────────────────────────
    # In override mode:  each corrected fraction is scaled so that the dry
    # components occupy (1 - water_frac_original) of the total.
    # In normal mode:    scale_factor = 1.0  (no change in behaviour).
    scale_factor = (1.0 - water_frac_original) if water_override else 1.0

    consolidated: list[ConsolidatedRow] = []
    total_hom = 0.0
    for rank, (g, frac_raw) in enumerate(included, 1):
        # frac_corrected within the DRY sub-stream, then scaled to total
        fc_dry  = frac_raw / frac_retained if frac_retained > 0 else 0.0
        fc      = fc_dry * scale_factor
        # mass computed relative to the dry sub-stream flows
        mass    = fc_dry * total_kmolh * g['phast_mw']
        total_hom += mass
        consolidated.append(ConsolidatedRow(
            rank               = rank,
            phast_name         = g['phast_name'],
            phast_mw           = g['phast_mw'],
            phast_status       = g['phast_status'],
            molar_frac_raw     = round(frac_raw, 8),
            molar_frac_pct_raw = round(frac_raw * 100, 6),
            frac_corrected     = round(fc, 8),
            frac_corrected_pct = round(fc * 100, 6),
            mass_kgh           = round(mass, 4),
            mass_lbh           = round(mass * LB_PER_KG, 4),
            mass_pct_w         = 0.0,
        ))

    # ── Append Water row at the END (override mode only) ─────────────────
    if water_override and water_frac_original > 0:
        water_compound = next(
            (c for c in all_compounds if c.phast_name == "Water"), None
        )
        w_phast_mw     = water_compound.mw     if water_compound else 18.015
        w_phast_status = water_compound.status if water_compound else 1
        # Approximate mass contribution of water to the homologated stream
        # We use the original stream total as reference base
        w_mass_kgh = water_frac_original * total_kmolh * w_phast_mw
        # Note: total_kmolh here is the DRY sub-stream; scale back to original
        # by dividing by (1 - water_frac_original) to recover original total
        if scale_factor > 0:
            orig_total_kmolh = total_kmolh / scale_factor
        else:
            orig_total_kmolh = total_kmolh
        w_mass_kgh = water_frac_original * orig_total_kmolh * w_phast_mw
        total_hom += w_mass_kgh

        water_rank = len(consolidated) + 1
        consolidated.append(ConsolidatedRow(
            rank               = water_rank,
            phast_name         = "Water",
            phast_mw           = w_phast_mw,
            phast_status       = w_phast_status,
            molar_frac_raw     = round(water_frac_original, 8),
            molar_frac_pct_raw = round(water_frac_original * 100, 6),
            frac_corrected     = round(water_frac_original, 8),
            frac_corrected_pct = round(water_frac_original * 100, 6),
            mass_kgh           = round(w_mass_kgh, 4),
            mass_lbh           = round(w_mass_kgh * LB_PER_KG, 4),
            mass_pct_w         = 0.0,
        ))

    # ── Mass percent (%w) ─────────────────────────────────────────────────
    for c in consolidated:
        c.mass_pct_w = round(c.mass_kgh / total_hom * 100, 4) if total_hom > 0 else 0.0

    molar_sum = round(sum(c.frac_corrected for c in consolidated), 8)

    return StreamResult(
        stream            = stream,
        homologated       = rows,
        consolidated      = consolidated,
        total_mass_kgh    = round(total_kgh, 4),
        total_molar_kmolh = round(total_kmolh, 6),
        n_requested       = n_requested,
        n_used            = len(consolidated),
        threshold         = round(threshold, 8),
        frac_retained     = round(frac_retained, 6),
        molar_sum_check   = molar_sum,
        water_override    = water_override,
        water_frac_original = water_frac_original,
    )


# ── Excel writer ───────────────────────────────────────────────────────────────

_NAVY="1F3864"; _ORANGE="E8720C"; _WHITE="FFFFFF"; _GRAY="F5F5F5"; _GN="E2EFDA"
_AM="FFF2CC";  _RED_LT="FCE4D6"; _WATER_BG="DDEEFF"; _WATER_FN="1F3864"

S_BG={1:"C6EFCE",2:"FFEB9C",3:"FFC7CE"}
S_FN={1:"276221",2:"9C5700",3:"9C0006"}
S_LB={1:"DISPONIBLE",2:"ADVERTENCIA",3:"PESADO"}

FLAG_BG  = {-2:"D9D9D9", -1:"FF0000",  0:"FF9900",  1:"FFFF00", None:"FFFFFF"}
FLAG_FN  = {-2:"404040", -1:"FFFFFF",  0:"FFFFFF",  1:"404040", None:"404040"}
FLAG_LBL = {-2:"INERT",  -1:"TOXIC",   0:"BOTH",    1:"FLAM",   None:"?"}

def _flag_lookup(all_compounds: list, phast_name: str) -> tuple:
    c = next((x for x in all_compounds if x.phast_name == phast_name), None)
    flag = c.flammable_toxic_flag if c else None
    return flag, FLAG_LBL.get(flag,"?"), FLAG_BG.get(flag,"FFFFFF"), FLAG_FN.get(flag,"000000")

def _flag_cell(ws, r, c, flag_int, label, bg, fg):
    x = ws.cell(r, c, label)
    x.fill = _fl(bg); x.font = _fn(bold=True, color=fg, sz=9)
    x.alignment = _al('center','center'); x.border = _bd()

def _fl(h): return PatternFill("solid", fgColor=h)
def _fn(bold=False, color="000000", sz=10):
    return Font(name="Calibri", bold=bold, color=color, size=sz)
def _bd():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h='left', v='center', w=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=w)

def _hc(ws, r, c, val, bg=_NAVY, fg=_WHITE, bold=True, h='center', sz=9, wrap=True):
    x = ws.cell(r, c, val)
    x.fill=_fl(bg); x.font=_fn(bold,fg,sz); x.alignment=_al(h,'center',wrap); x.border=_bd()
    return x

def _dc(ws, r, c, val, fmt=None, bg=_WHITE, bold=False, h='right'):
    x = ws.cell(r, c, val)
    x.fill=_fl(bg); x.font=_fn(bold); x.alignment=_al(h,'center'); x.border=_bd()
    if fmt: x.number_format = fmt
    return x

def _sc(ws, r, c, status):
    x = ws.cell(r, c, S_LB.get(status,'?'))
    x.fill=_fl(S_BG.get(status,_WHITE))
    x.font=_fn(True, S_FN.get(status,"000000"), 9)
    x.alignment=_al('center','center'); x.border=_bd()


def _water_cell(ws, r, c_start, n_cols, label="WATER"):
    """Paint a highlighted water-override indicator across n_cols columns."""
    ws.merge_cells(start_row=r, start_column=c_start,
                   end_row=r,   end_column=c_start + n_cols - 1)
    x = ws.cell(r, c_start, label)
    x.fill=_fl(_WATER_BG); x.font=_fn(bold=True, color=_WATER_FN, sz=9)
    x.alignment=_al('center','center'); x.border=_bd()


def _safe_sheet_title(raw: str, max_len: int = 28) -> str:
    forbidden = set('[]' + ':*?/\\')
    cleaned = ''.join(c for c in raw if c not in forbidden)
    return cleaned[:max_len].strip()


def write_output(
    results_input,
    out_path: str | Path,
    N: int,
    all_compounds: list | None = None,
    water_threshold: float = WATER_OVERRIDE_DEFAULT,
):
    """
    Write Excel output.  Accepts:
      - list[StreamResult]
      - list[tuple[str, StreamResult]]

    Water-override rows are highlighted with a pale-blue background in
    the consolidated table, and the Water row is always last with a
    distinct colour band.
    """
    if results_input and isinstance(results_input[0], tuple):
        labeled: list[tuple[str, StreamResult]] = results_input
    else:
        labeled = [("BASE", r) for r in results_input]

    OPT_BG  = {"BASE":"DEEAF1","OPT1_DRY":"E2EFDA","OPT2_REM":"FFF2CC","OPT3_FORCE":"FCE4D6"}
    OPT_FN  = {"BASE":"1F3864","OPT1_DRY":"375623","OPT2_REM":"7F6000","OPT3_FORCE":"9C0006"}
    OPT_LBL = {"BASE":"BASE","OPT1_DRY":"BASE SECA","OPT2_REM":"SIN HC","OPT3_FORCE":"FORZADO"}

    wb = Workbook()

    # ── Sheet 1: Global summary ────────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Resumen_Homologacion"; ws0.freeze_panes = "G3"
    h0 = [
        "Opción","Proyecto","Escenario","Corriente",
        "Water Override","Umbral agua (%)",
        "Total masa (kg/h)","Total molar (kgmol/h)",
        "N solic.","N result.","% Retenido",
        "Rank","Sustancia PHAST","PM (g/mol)","Status","Flag Peligro",
        "% Mol (raw)","% Mol corregido (Σ=100%)","Flujo (kg/h)","Flujo (lb/h)","% Másico",
    ]
    w0 = [11,14,10,18,14,14,20,20,8,8,12,6,28,12,14,11,13,20,18,18,12]
    for i,(h,w) in enumerate(zip(h0,w0),1):
        _hc(ws0,1,i,h); ws0.column_dimensions[get_column_letter(i)].width=w
    ws0.row_dimensions[1].height=45

    row=2
    for label, res in labeled:
        s=res.stream; pr=f"{res.frac_retained*100:.2f}%"
        opt_bg = OPT_BG.get(label, "FFFFFF")
        opt_fn = OPT_FN.get(label, "000000")
        opt_lb = OPT_LBL.get(label, label)
        wo_label = f"SÍ (≥{water_threshold*100:.0f}%)" if res.water_override else "No"
        wo_bg    = _WATER_BG if res.water_override else _WHITE
        for ri, c in enumerate(res.consolidated):
            # Highlight water row differently
            is_water_row = (res.water_override and c.phast_name == "Water"
                            and ri == len(res.consolidated) - 1)
            bg = _WATER_BG if is_water_row else (_GRAY if ri % 2 else _WHITE)

            oc = ws0.cell(row,1,opt_lb)
            oc.fill=_fl(opt_bg); oc.font=_fn(bold=True,color=opt_fn,sz=9)
            oc.alignment=_al('center','center'); oc.border=_bd()
            _dc(ws0,row,2,str(s.project),bg=bg,h='left')
            _dc(ws0,row,3,s.escenario,bg=bg,h='center')
            _dc(ws0,row,4,str(s.corriente),bg=bg,h='center')
            # Water override indicator columns (5,6)
            wc = ws0.cell(row,5,wo_label)
            wc.fill=_fl(wo_bg); wc.font=_fn(bold=res.water_override,color=_WATER_FN if res.water_override else "404040",sz=9)
            wc.alignment=_al('center','center'); wc.border=_bd()
            _dc(ws0,row,6,f"{water_threshold*100:.0f}%" if res.water_override else "—",bg=wo_bg,h='center')
            _dc(ws0,row,7,res.total_mass_kgh,bg=bg,fmt='#,##0.0000')
            _dc(ws0,row,8,res.total_molar_kmolh,bg=bg,fmt='#,##0.0000')
            _dc(ws0,row,9,res.n_requested,bg=bg,h='center')
            _dc(ws0,row,10,res.n_used,bg=bg,h='center')
            _dc(ws0,row,11,pr,bg=bg,h='center')
            _dc(ws0,row,12,c.rank,bg=bg,h='center')
            _dc(ws0,row,13,c.phast_name,bg=bg,h='left',bold=True)
            _dc(ws0,row,14,c.phast_mw,bg=bg,fmt='0.000')
            _sc(ws0,row,15,c.phast_status)
            flag_i, flag_lbl, flag_bg, flag_fg = _flag_lookup(all_compounds or [], c.phast_name)
            _flag_cell(ws0,row,16,flag_i,flag_lbl,flag_bg,flag_fg)
            _dc(ws0,row,17,c.molar_frac_pct_raw,bg=bg,fmt='0.0000')
            _dc(ws0,row,18,c.frac_corrected_pct,bg=bg,fmt='0.0000')
            _dc(ws0,row,19,c.mass_kgh,bg=bg,fmt='#,##0.0000')
            _dc(ws0,row,20,c.mass_lbh,bg=bg,fmt='#,##0.0000')
            _dc(ws0,row,21,c.mass_pct_w,bg=bg,fmt='0.0000')
            row+=1

    # ── Sheets 2+: Per-stream detail ──────────────────────────────────────
    used: set[str] = set()
    for label, res in labeled:
        s = res.stream
        opt_suffix = {"BASE":"","OPT1_DRY":"_DRY","OPT2_REM":"_REM","OPT3_FORCE":"_FORCE"}.get(label,"")
        corr_clean = str(s.corriente).split("[")[0].strip()
        raw_title  = f"Esc{s.escenario}_C{corr_clean}{opt_suffix}"
        base  = _safe_sheet_title(raw_title, 28)
        title = base; n=1
        while title in used: title=f"{base[:25]}_{n}"; n+=1
        used.add(title)
        ws = wb.create_sheet(title); ws.freeze_panes="A3"

        opt_banner_bg = OPT_BG.get(label, _NAVY)
        opt_banner_fn = OPT_FN.get(label, _WHITE)
        override_note = f"  ★ WATER OVERRIDE ACTIVO (fracción H₂O original = {res.water_frac_original*100:.2f}% molar)" if res.water_override else ""
        ws.merge_cells("A1:K1")
        hc=ws.cell(1,1,(
            f"[{OPT_LBL.get(label,label)}]  HOMOLOGACIÓN  ·  Proyecto: {s.project}  ·  "
            f"Esc: {s.escenario}  ·  Corriente: {s.corriente}  ·  "
            f"Entrada: {res.total_mass_kgh:,.2f} kg/h ({res.total_molar_kmolh:,.4f} kgmol/h)  ·  "
            f"N={res.n_used}  ·  Retenido: {res.frac_retained*100:.2f}% → renorm. 100%"
            f"{override_note}"
        ))
        hc.fill=_fl(opt_banner_bg)
        hc.font=Font(name="Calibri",bold=True,color=opt_banner_fn,size=10)
        hc.alignment=_al('left','center'); ws.row_dimensions[1].height=24

        # ── Part A: Homologation detail ────────────────────────────────────
        hdrs_a = ["Componente","Flujo másico (kg/h)","Flujo molar (kgmol/h)",
                  "PM calc. (g/mol)","Sustancia PHAST","PM PHAST (g/mol)",
                  "Flag Peligro","Frac. molar","Pool","Método","Error PM%","Advertencia"]
        for ci,h in enumerate(hdrs_a,1):
            _hc(ws,2,ci,h,bg=_ORANGE)
        ws.row_dimensions[2].height=30

        for ri,hr in enumerate(res.homologated):
            r3=ri+3; bg=_GRAY if ri%2 else _WHITE
            _dc(ws,r3,1,hr.source_name,bg=bg,h='left')
            _dc(ws,r3,2,hr.mass_kgh,bg=bg,fmt='#,##0.0000')
            _dc(ws,r3,3,hr.molar_kmolh,bg=bg,fmt='#,##0.000000')
            _dc(ws,r3,4,hr.mw_calc,bg=bg,fmt='0.00')
            _dc(ws,r3,5,hr.phast_name,bg=bg,h='left',bold=True)
            _dc(ws,r3,6,hr.phast_mw,bg=bg,fmt='0.000')
            fi,fl,fb,ff = _flag_lookup(all_compounds or [],hr.phast_name)
            _flag_cell(ws,r3,7,fi,fl,fb,ff)
            _dc(ws,r3,8,hr.molar_frac,bg=bg,fmt='0.00000000')
            pool_bg  = "DEEBF7" if hr.match_pool=="hydrocarbon_only" else _WHITE
            pool_lbl = "HC-ONLY" if hr.match_pool=="hydrocarbon_only" else "FULL"
            pool_clr = "1F4E79" if hr.match_pool=="hydrocarbon_only" else "404040"
            pc = ws.cell(r3,9,pool_lbl)
            pc.fill=_fl(pool_bg); pc.font=_fn(True,pool_clr,8)
            pc.alignment=_al('center','center'); pc.border=_bd()
            method_lbl = {"alias":"ALIAS","mw_unique":"MW","mw_tie_resolved_by_alias":"TIE+ALIAS",
                          "mw_tie_unresolved":"TIE","fallback_no_mw":"FALLBACK"}.get(hr.match_method,hr.match_method)
            mc2 = ws.cell(r3,10,method_lbl)
            mc2.fill=_fl(bg); mc2.font=_fn(sz=8); mc2.alignment=_al('center','center'); mc2.border=_bd()
            err_bg = _RED_LT if hr.mw_error_pct > MW_WARN_PCT else bg
            _dc(ws,r3,11,hr.mw_error_pct,bg=err_bg,fmt='0.00')
            _dc(ws,r3,12,hr.warning,bg=err_bg if hr.warning else bg,h='left')

        col_widths_a=[22,16,18,14,28,14,10,16,9,12,10,35]
        for ci,w in enumerate(col_widths_a,1):
            ws.column_dimensions[get_column_letter(ci)].width=w

        # ── Water override note in Part A (if applicable) ──────────────────
        if res.water_override:
            note_row = len(res.homologated) + 3
            ws.merge_cells(f"A{note_row}:L{note_row}")
            nc = ws.cell(note_row, 1,
                f"★ WATER OVERRIDE ACTIVO — Agua excluida del balance de homologación "
                f"(fracción molar original = {res.water_frac_original*100:.4f}%). "
                f"Se reintegra al final del resultado con fracción intacta."
            )
            nc.fill=_fl(_WATER_BG); nc.font=_fn(bold=True,color=_WATER_FN,sz=9)
            nc.alignment=_al('left','center',True); ws.row_dimensions[note_row].height=30

        sep = len(res.homologated) + (4 if res.water_override else 3)
        ws.row_dimensions[sep].height=8

        # ── Part B: Consolidated result ────────────────────────────────────
        tr=sep+1; ws.merge_cells(f"A{tr}:K{tr}")
        override_note2 = (
            f"  WATER: fracción H₂O = {res.water_frac_original*100:.4f}% (fija) · "
            f"resto renorm. en espacio ({(1-res.water_frac_original)*100:.4f}%)"
            if res.water_override else ""
        )
        tc=ws.cell(tr,1,(
            f"RESULTADO CONSOLIDADO  ·  N={res.n_used}  ·  "
            f"Umbral: {res.threshold:.6f}  ·  "
            f"Retenido: {res.frac_retained*100:.2f}%  ·  "
            f"Renorm. → Σ%mol=100%  Σ%w=100%"
            f"{override_note2}"
        ))
        tc.fill=_fl(_NAVY); tc.font=Font(name="Calibri",bold=True,color=_WHITE,size=10)
        tc.alignment=_al('left','center'); ws.row_dimensions[tr].height=22

        hrr=tr+1
        hdrs_b=["Rank","Sustancia PHAST","PM (g/mol)","Status","Flag Peligro",
                "% Mol (raw)","% Mol corregido (Σ=100%)","Frac. corregida (Σ=1)",
                "Flujo (kg/h)","Flujo (lb/h)","% Másico (Σ=100%)"]
        for ci,h in enumerate(hdrs_b,1):
            _hc(ws,hrr,ci,h,bg=_NAVY)
        ws.row_dimensions[hrr].height=40

        for ri,c in enumerate(res.consolidated):
            rb=hrr+1+ri
            is_water_row = (res.water_override and c.phast_name == "Water"
                            and ri == len(res.consolidated) - 1)
            bg = _WATER_BG if is_water_row else (_GRAY if ri%2 else _WHITE)
            _dc(ws,rb,1,c.rank,bg=bg,h='center')
            name_val = f"★ {c.phast_name} [OVERRIDE]" if is_water_row else c.phast_name
            _dc(ws,rb,2,name_val,bg=bg,h='left',bold=True)
            _dc(ws,rb,3,c.phast_mw,bg=bg,fmt='0.000')
            _sc(ws,rb,4,c.phast_status)
            fi,fl,fb,ff = _flag_lookup(all_compounds or [],c.phast_name)
            _flag_cell(ws,rb,5,fi,fl,fb,ff)
            _dc(ws,rb,6,c.molar_frac_pct_raw,bg=bg,fmt='0.0000')
            _dc(ws,rb,7,c.frac_corrected_pct,bg=bg,fmt='0.0000')
            _dc(ws,rb,8,c.frac_corrected,bg=bg,fmt='0.00000000')
            _dc(ws,rb,9,c.mass_kgh,bg=bg,fmt='#,##0.0000')
            _dc(ws,rb,10,c.mass_lbh,bg=bg,fmt='#,##0.0000')
            _dc(ws,rb,11,c.mass_pct_w,bg=bg,fmt='0.0000')

        tot=hrr+len(res.consolidated)+1
        _hc(ws,tot,1,"TOTAL",bg=_ORANGE,bold=True)
        for ci in range(2,6): _hc(ws,tot,ci,"",bg=_ORANGE)
        _dc(ws,tot,6,round(sum(c.molar_frac_pct_raw  for c in res.consolidated),4),fmt='0.0000',bg=_GN,bold=True)
        _dc(ws,tot,7,round(sum(c.frac_corrected_pct  for c in res.consolidated),4),fmt='0.0000',bg=_GN,bold=True)
        _dc(ws,tot,8,res.molar_sum_check,fmt='0.00000000',bg=_GN,bold=True)
        _dc(ws,tot,9,round(sum(c.mass_kgh  for c in res.consolidated),4),fmt='#,##0.0000',bg=_GN,bold=True)
        _dc(ws,tot,10,round(sum(c.mass_lbh  for c in res.consolidated),4),fmt='#,##0.0000',bg=_GN,bold=True)
        _dc(ws,tot,11,round(sum(c.mass_pct_w for c in res.consolidated),4),fmt='0.0000',bg=_GN,bold=True)

        for ci,w in enumerate([6,28,12,14,10,13,22,20,18,18,14],1):
            ws.column_dimensions[get_column_letter(ci)].width=w

    wb.save(out_path)
    print(f"\n✓ Output saved: {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# POST-HOMOLOGATION OPTIONS (unchanged from v1)
# ══════════════════════════════════════════════════════════════════════════════

def _renormalise(
    rows: list[tuple[dict,float]],
    corrected_scale: float = 1.0,
    fixed_water: ConsolidatedRow | None = None,
) -> list[ConsolidatedRow]:
    total_frac = sum(f for _,f in rows)
    consolidated = []
    total_hom = 0.0
    for rank,(g,frac_raw) in enumerate(rows,1):
        fc_norm = frac_raw / total_frac if total_frac > 0 else 0.0
        fc      = fc_norm * corrected_scale
        mass    = fc_norm * g.get('total_kmolh', 1.0) * g['phast_mw']
        total_hom += mass
        consolidated.append(ConsolidatedRow(
            rank=rank, phast_name=g['phast_name'],
            phast_mw=g['phast_mw'], phast_status=g['phast_status'],
            molar_frac_raw=round(frac_raw,8),
            molar_frac_pct_raw=round(frac_raw*100,6),
            frac_corrected=round(fc,8),
            frac_corrected_pct=round(fc*100,6),
            mass_kgh=round(mass,4), mass_lbh=round(mass*LB_PER_KG,4),
            mass_pct_w=0.0))
    if fixed_water is not None:
        water = ConsolidatedRow(
            rank=len(consolidated) + 1,
            phast_name=fixed_water.phast_name,
            phast_mw=fixed_water.phast_mw,
            phast_status=fixed_water.phast_status,
            molar_frac_raw=fixed_water.molar_frac_raw,
            molar_frac_pct_raw=fixed_water.molar_frac_pct_raw,
            frac_corrected=fixed_water.frac_corrected,
            frac_corrected_pct=fixed_water.frac_corrected_pct,
            mass_kgh=fixed_water.mass_kgh,
            mass_lbh=fixed_water.mass_lbh,
            mass_pct_w=0.0,
        )
        total_hom += water.mass_kgh
        consolidated.append(water)
    for c in consolidated:
        c.mass_pct_w = round(c.mass_kgh/total_hom*100,4) if total_hom > 0 else 0.0
    return consolidated


def _print_table(headers: list[str], rows: list[list[str]], col_widths: list[int]):
    fmt = "  " + "  ".join(f"{{:<{w}}}" for w in col_widths)
    sep = "  " + "  ".join("─" * w for w in col_widths)
    print(fmt.format(*headers))
    print(sep)
    for row in rows:
        print(fmt.format(*[str(v) for v in row]))


def apply_post_options(
    streams:          list[StreamInput],
    results:          list[StreamResult],
    all_compounds:    list[PhastCompound],
    hydrocarbon_only: list[PhastCompound],
    aliases:          dict[str, str],
) -> list[tuple[str, StreamResult]]:
    flag_map = {c.phast_name: c for c in all_compounds}

    flagged = []
    for idx, (stream, res) in enumerate(zip(streams, results)):
        all_hom   = {hr.phast_name for hr in res.homologated}
        in_result = {cr.phast_name for cr in res.consolidated}
        missing   = [n for n in all_hom
                     if flag_map.get(n) and flag_map[n].is_toxic
                     and n not in in_result]
        flagged.append((idx, stream, res, missing))

    streams_with_missing = [(i,s,r,m) for i,s,r,m in flagged if m]

    if not streams_with_missing:
        print("\n✓ Todas las corrientes incluyen sus compuestos tóxicos/ambos en el resultado.")
        return [("BASE", r) for r in results]

    print()
    print("═" * 72)
    print("  ANÁLISIS POST-HOMOLOGACIÓN — Compuestos tóxicos ausentes del resultado")
    print("═" * 72)
    for pos, (i, stream, res, missing) in enumerate(streams_with_missing, 1):
        print(f"\n  [{pos}] Escenario {stream.escenario}  |  Corriente {stream.corriente}")
        print(f"       Ausentes: {', '.join(missing)}")

    print()
    raw = input(
        "  ¿Sobre qué corrientes aplicar opciones?\n"
        "  (números de la lista, ej: 1,3  |  Enter = todas): "
    ).strip()
    if raw:
        try:
            sel_idx = [int(x.strip()) - 1 for x in raw.split(",")]
            selected = [streams_with_missing[i] for i in sel_idx
                        if 0 <= i < len(streams_with_missing)]
        except Exception:
            selected = streams_with_missing
    else:
        selected = streams_with_missing

    output_map: dict[int, list[tuple[str, StreamResult]]] = {}
    for i, res in enumerate(results):
        output_map[i] = [("BASE", res)]

    for i, stream, res, missing in selected:
        print()
        print(f"╔{'═'*68}╗")
        print(f"║  Escenario {stream.escenario}  |  Corriente {stream.corriente:<48}║")
        print(f"║  Tóxicos ausentes: {', '.join(missing):<49}║")
        print(f"╚{'═'*68}╝")

        has_water = any(
            c.name.strip().lower() in WATER_KEYS
            for c in stream.components
        )
        water_frac = sum(
            c.molar_kmolh for c in stream.components
            if c.name.strip().lower() in WATER_KEYS
        ) / max(res.total_molar_kmolh, 1e-9)

        has_pure_hc = any(
            flag_map.get(cr.phast_name) and flag_map[cr.phast_name].is_flammable
            and not flag_map[cr.phast_name].is_toxic
            for cr in res.consolidated
        )

        print()
        print("  Opciones disponibles (pueden combinarse — cada una genera resultado independiente):")
        opt1_note = f"  [agua = {water_frac*100:.1f}% molar]" if has_water else "  [sin agua detectada]"
        print(f"    1 → Base seca: retirar H₂O y re-homologar {opt1_note}")
        print(f"    2 → Remover hidrocarburos: liberar slots para tóxicos"
              + ("" if has_pure_hc else "  [no hay HC puros en resultado]"))
        print(f"    3 → Forzar tóxicos: incluir compuesto(s) específicos sí o sí")
        print(f"    0 → Sin cambios para esta corriente")
        print()

        opts_raw = input("  Selecciona opciones (ej: 1,3  |  1,2,3  |  0): ").strip()
        if not opts_raw or opts_raw == "0":
            print("  Sin cambios.")
            continue

        try:
            opts_chosen = [int(x.strip()) for x in opts_raw.split(",")
                           if x.strip() in ("1","2","3")]
        except Exception:
            print("  Entrada inválida — sin cambios.")
            continue

        opts_chosen = sorted(set(opts_chosen))

        for opt in opts_chosen:
            print()
            if opt == 1:
                print("  ── OPCIÓN 1: Base seca ──────────────────────────────────────")
                n_raw = input(
                    f"  N para base seca [1-{PHAST_MAX}, default={res.n_requested}]: "
                ).strip()
                n_new = int(n_raw) if n_raw.isdigit() and 1 <= int(n_raw) <= PHAST_MAX \
                        else res.n_requested

                dry_comps = [c for c in stream.components
                             if c.name.strip().lower() not in WATER_KEYS]
                water_mass  = sum(c.mass_kgh    for c in stream.components
                                  if c.name.strip().lower() in WATER_KEYS)
                water_molar = sum(c.molar_kmolh for c in stream.components
                                  if c.name.strip().lower() in WATER_KEYS)

                if not dry_comps:
                    print("  ⚠ No quedan componentes sin agua — opción omitida.")
                    continue

                print(f"  Retirando H₂O: {water_mass:,.2f} kg/h  ({water_molar:,.4f} kgmol/h)")
                print(f"  Re-homologando en base seca con N={n_new}...")

                dry_stream = StreamInput(
                    project=stream.project, escenario=stream.escenario,
                    corriente=f"{stream.corriente} [BASE SECA]",
                    mass_unit=stream.mass_unit, mol_unit=stream.mol_unit,
                    components=dry_comps,
                )
                alt_res = homologate(dry_stream, all_compounds, hydrocarbon_only, aliases, n_new)
                smol = sum(c.frac_corrected_pct for c in alt_res.consolidated)
                sw   = sum(c.mass_pct_w         for c in alt_res.consolidated)
                print(f"  ✓ {alt_res.n_used} compuestos  Σ%mol={smol:.4f}%  Σ%w={sw:.4f}%")
                output_map[i].append(("OPT1_DRY", alt_res))

            elif opt == 2:
                print("  ── OPCIÓN 2: Remover hidrocarburos puros del resultado ──────")
                hc_in_result = [
                    cr for cr in res.consolidated
                    if flag_map.get(cr.phast_name)
                    and flag_map[cr.phast_name].is_flammable
                    and not flag_map[cr.phast_name].is_toxic
                ]
                if not hc_in_result:
                    print("  ⚠ No hay hidrocarburos puros (flag=1) en el resultado base — opción omitida.")
                    continue

                hc_sorted = sorted(hc_in_result, key=lambda c: c.mass_kgh)
                print()
                headers = ["No.", "Compuesto PHAST", "% Molar", "% Másico", "Flujo másico (kg/h)"]
                widths  = [4, 32, 10, 10, 20]
                rows_t  = [
                    [str(j+1), c.phast_name,
                     f"{c.frac_corrected_pct:.4f}",
                     f"{c.mass_pct_w:.4f}",
                     f"{c.mass_kgh:,.4f}"]
                    for j, c in enumerate(hc_sorted)
                ]
                _print_table(headers, rows_t, widths)
                print()

                raw_sel = input("  Números de HC a remover (ej: 1,3  |  Enter = cancelar): ").strip()
                if not raw_sel:
                    print("  Sin cambios para Opción 2.")
                    continue

                try:
                    sel_nos = [int(x.strip()) - 1 for x in raw_sel.split(",")]
                    remove_names = {hc_sorted[j].phast_name for j in sel_nos
                                    if 0 <= j < len(hc_sorted)}
                except Exception:
                    print("  Entrada inválida — Opción 2 omitida.")
                    continue

                print(f"  Removiendo: {', '.join(remove_names)}")
                fixed_water = (
                    next((c for c in res.consolidated if c.phast_name == "Water"), None)
                    if res.water_override else None
                )
                kept = [(c, c.molar_frac_raw)
                        for c in res.consolidated
                        if c.phast_name not in remove_names
                        and not (fixed_water is not None and c.phast_name == "Water")]
                if not kept:
                    print("  ⚠ No quedan compuestos — Opción 2 omitida.")
                    continue

                tk = res.total_molar_kmolh
                rows_rn = [({'phast_name':c.phast_name,'phast_mw':c.phast_mw,
                              'phast_status':c.phast_status,'total_kmolh':tk}, f)
                           for c,f in kept]
                dry_scale = (1.0 - res.water_frac_original) if fixed_water else 1.0
                new_cons = _renormalise(rows_rn, dry_scale, fixed_water)

                alt_res = StreamResult(
                    stream=StreamInput(
                        project=stream.project, escenario=stream.escenario,
                        corriente=f"{stream.corriente} [SIN HC]",
                        mass_unit=stream.mass_unit, mol_unit=stream.mol_unit,
                        components=stream.components,
                    ),
                    homologated=res.homologated,
                    consolidated=new_cons,
                    total_mass_kgh=res.total_mass_kgh,
                    total_molar_kmolh=res.total_molar_kmolh,
                    n_requested=res.n_requested,
                    n_used=len(new_cons),
                    threshold=res.threshold,
                    frac_retained=round(sum(f for _,f in kept), 6),
                    molar_sum_check=round(sum(c.frac_corrected for c in new_cons), 8),
                    water_override=res.water_override,
                    water_frac_original=res.water_frac_original,
                )
                smol = sum(c.frac_corrected_pct for c in alt_res.consolidated)
                sw   = sum(c.mass_pct_w         for c in alt_res.consolidated)
                print(f"  ✓ {alt_res.n_used} compuestos  Σ%mol={smol:.4f}%  Σ%w={sw:.4f}%")
                output_map[i].append(("OPT2_REM", alt_res))

            elif opt == 3:
                print("  ── OPCIÓN 3: Forzar compuestos tóxicos específicos ─────────")
                all_hom_names  = {hr.phast_name for hr in res.homologated}
                in_base_result = {cr.phast_name for cr in res.consolidated}

                toxic_cands = [
                    c for c in all_compounds
                    if c.phast_name in all_hom_names and c.is_toxic
                ]
                if not toxic_cands:
                    print("  ⚠ No hay compuestos tóxicos/ambos en el balance de entrada — opción omitida.")
                    continue

                print()
                headers = ["No.", "Compuesto PHAST", "Flag", "%mol en entrada", "En resultado base"]
                widths  = [4, 35, 6, 17, 17]
                rows_t  = []
                for j, c in enumerate(toxic_cands):
                    frac_in = sum(hr.molar_frac for hr in res.homologated
                                  if hr.phast_name == c.phast_name)
                    status  = "SÍ ✓" if c.phast_name in in_base_result else "NO ✗"
                    flag_s  = FLAG_LBL.get(c.flammable_toxic_flag, "?")
                    rows_t.append([str(j+1), c.phast_name, flag_s,
                                   f"{frac_in*100:.4f}%", status])
                _print_table(headers, rows_t, widths)
                print()

                raw_f = input("  Números de compuestos a forzar (ej: 1,2  |  Enter = cancelar): ").strip()
                if not raw_f:
                    print("  Sin cambios para Opción 3.")
                    continue

                try:
                    fidxs = [int(x.strip()) - 1 for x in raw_f.split(",")]
                    force_names = [toxic_cands[j].phast_name for j in fidxs
                                   if 0 <= j < len(toxic_cands)]
                except Exception:
                    print("  Entrada inválida — Opción 3 omitida.")
                    continue

                print(f"  Forzando: {', '.join(force_names)}")

                tk = res.total_molar_kmolh
                fixed_water = (
                    next((c for c in res.consolidated if c.phast_name == "Water"), None)
                    if res.water_override else None
                )
                compound_by_name = {c.phast_name: c for c in all_compounds}
                forced_c = [compound_by_name[n] for n in force_names if n in compound_by_name]
                dry_slots_total = max(0, res.n_requested - (1 if fixed_water else 0))
                if len(forced_c) > dry_slots_total:
                    print(f"  ⚠ Agua ocupa 1 cupo; se forzarán solo {dry_slots_total} compuesto(s) seco(s).")
                    forced_c = forced_c[:dry_slots_total]
                forced_rows_rn = []
                for c in forced_c:
                    frac = sum(hr.molar_frac for hr in res.homologated
                               if hr.phast_name == c.phast_name)
                    forced_rows_rn.append(
                        ({'phast_name':c.phast_name,'phast_mw':c.mw,
                          'phast_status':c.status,'total_kmolh':tk}, frac)
                    )

                remaining = [(c, c.molar_frac_raw)
                             for c in res.consolidated
                             if c.phast_name not in set(force_names)
                             and not (fixed_water is not None and c.phast_name == "Water")]
                slots = max(0, dry_slots_total - len(forced_rows_rn))
                fill_rows_rn = [
                    ({'phast_name':c.phast_name,'phast_mw':c.phast_mw,
                      'phast_status':c.phast_status,'total_kmolh':tk}, f)
                    for c,f in remaining[:slots]
                ]

                all_rows_rn = forced_rows_rn + fill_rows_rn
                if not all_rows_rn:
                    print("  ⚠ Sin compuestos — Opción 3 omitida.")
                    continue

                dry_scale = (1.0 - res.water_frac_original) if fixed_water else 1.0
                new_cons = _renormalise(all_rows_rn, dry_scale, fixed_water)

                alt_res = StreamResult(
                    stream=StreamInput(
                        project=stream.project, escenario=stream.escenario,
                        corriente=f"{stream.corriente} [FORZADO]",
                        mass_unit=stream.mass_unit, mol_unit=stream.mol_unit,
                        components=stream.components,
                    ),
                    homologated=res.homologated,
                    consolidated=new_cons,
                    total_mass_kgh=res.total_mass_kgh,
                    total_molar_kmolh=res.total_molar_kmolh,
                    n_requested=res.n_requested,
                    n_used=len(new_cons),
                    threshold=res.threshold,
                    frac_retained=round(sum(f for _,f in all_rows_rn), 6),
                    molar_sum_check=round(sum(c.frac_corrected for c in new_cons), 8),
                    water_override=res.water_override,
                    water_frac_original=res.water_frac_original,
                )
                smol = sum(c.frac_corrected_pct for c in alt_res.consolidated)
                sw   = sum(c.mass_pct_w         for c in alt_res.consolidated)
                print(f"  ✓ {alt_res.n_used} compuestos  Σ%mol={smol:.4f}%  Σ%w={sw:.4f}%")
                output_map[i].append(("OPT3_FORCE", alt_res))

    final: list[tuple[str, StreamResult]] = []
    for i in range(len(results)):
        final.extend(output_map.get(i, [("BASE", results[i])]))
    return final


# ── Interactive helpers ────────────────────────────────────────────────────────

def prompt_int(msg: str, lo: int, hi: int, default: int) -> int:
    while True:
        raw = input(f"  {msg} [{lo}-{hi}, default={default}]: ").strip()
        if raw == '':
            return default
        try:
            v = int(raw)
            if lo <= v <= hi:
                return v
        except ValueError:
            pass
        print(f"    → Por favor ingresa un número entre {lo} y {hi}.")


def prompt_float(msg: str, lo: float, hi: float, default: float) -> float:
    """Ask for a float in [lo, hi], return default on blank input."""
    while True:
        raw = input(f"  {msg} [{lo:.0%}-{hi:.0%}, default={default:.0%}]: ").strip()
        if raw == '':
            return default
        # Accept either "0.85" or "85" or "85%"
        cleaned = raw.replace('%', '').strip()
        try:
            v = float(cleaned)
            if v > 1.0:   # assume percentage entry e.g. "85"
                v /= 100.0
            if lo <= v <= hi:
                return v
        except ValueError:
            pass
        print(f"    → Ingresa un valor entre {lo:.0%} y {hi:.0%} (ej: 0.80 ó 80).")


def prompt_file(label: str, extensions: tuple[str, ...], must_exist: bool = True) -> Path:
    script_dir = Path(__file__).parent
    ext_str = " / ".join(extensions)
    while True:
        raw = input(f"  Ruta del archivo {label} ({ext_str}): ").strip().strip('"').strip("'")
        if not raw:
            print("    → La ruta no puede estar vacía.")
            continue
        p = Path(raw)
        if not p.is_absolute():
            p = script_dir / p
        if must_exist and not p.exists():
            print(f"    → Archivo no encontrado: {p}")
            continue
        if not any(p.suffix.lower() == ext for ext in extensions):
            print(f"    → Extensión inesperada '{p.suffix}'. Se esperaba: {ext_str}")
        return p


def prompt_output(label: str, default_name: str) -> Path:
    script_dir = Path(__file__).parent
    default_path = script_dir / default_name
    raw = input(f"  Ruta de salida [{default_path}]: ").strip().strip('"').strip("'")
    if not raw:
        return default_path
    p = Path(raw)
    if not p.is_absolute():
        p = script_dir / p
    return p


# ── Main runner ───────────────────────────────────────────────────────────────

def run(input_path, db_path, aliases_path, output_path,
        n_compounds=None, water_threshold=None):
    print("=" * 60)
    print("  PHAST Homologator v2 — phast_pseudomps")
    print("=" * 60)

    print("\nCargando base de datos PHAST...")
    all_compounds, hydrocarbon_only = load_db(db_path)
    print(f"  Compuestos totales:        {len(all_compounds)}")
    print(f"  Pool hidrocarbonos (CnHm): {len(hydrocarbon_only)}")

    print("\nCargando tabla de alias...")
    aliases = load_aliases(aliases_path)
    print(f"  Alias cargados:            {len(aliases)} entradas (built-in + externos)")

    print("\nLeyendo plantilla de entrada...")
    streams = parse_input(input_path)
    print(f"  Corrientes encontradas: {len(streams)}")
    for s in streams:
        print(f"    Esc={s.escenario}  C={s.corriente}  → {len(s.components)} componentes")

    # ── Water threshold ────────────────────────────────────────────────────
    if water_threshold is None:
        print()
        print("── Umbral de Water Override ─────────────────────────────────")
        print("  Corrientes con fracción molar de agua ≥ umbral recibirán")
        print("  el override: el agua pasa con fracción original sin normalizar.")
        water_threshold = prompt_float(
            "Umbral de agua (fracción molar)", 0.01, 0.99, WATER_OVERRIDE_DEFAULT
        )
    print(f"\n  Umbral Water Override: {water_threshold:.1%} molar")

    # ── N compounds ────────────────────────────────────────────────────────
    if n_compounds is None:
        print()
        print(f"¿Cuántos compuestos PHAST conservar por corriente? (máximo {PHAST_MAX})")
        n_compounds = prompt_int("N compuestos", 1, PHAST_MAX, 14)

    # ── ★ Water Override pre-scan ──────────────────────────────────────────
    print("\n── Escaneando contenido de agua en corrientes ───────────────")
    scan_results = scan_water_content(streams)
    override_indices = prompt_water_override(scan_results, water_threshold)

    # ── Homologation loop ──────────────────────────────────────────────────
    print(f"\nEjecutando homologación  N={n_compounds}...")
    results: list[StreamResult] = []

    for idx, s in enumerate(streams):
        apply_override = idx in override_indices

        if apply_override:
            # Separate water from the stream
            water_kmolh = sum(c.molar_kmolh for c in s.components if _is_water(c.name))
            total_kmolh = sum(c.molar_kmolh for c in s.components)
            water_frac  = water_kmolh / total_kmolh if total_kmolh > 0 else 0.0

            dry_comps = [c for c in s.components if not _is_water(c.name)]
            dry_stream = StreamInput(
                project=s.project, escenario=s.escenario,
                corriente=s.corriente,          # keep original corriente name
                mass_unit=s.mass_unit, mol_unit=s.mol_unit,
                components=dry_comps,
            )
            print(f"  [WATER OVERRIDE]  Esc={s.escenario}  C={str(s.corriente):<10}  "
                  f"x_agua={water_frac*100:.2f}% → homologando base seca...")
            res = homologate(
                dry_stream, all_compounds, hydrocarbon_only, aliases, n_compounds,
                water_override=True, water_frac_original=water_frac,
            )
        else:
            res = homologate(s, all_compounds, hydrocarbon_only, aliases, n_compounds)

        results.append(res)
        pseudo_count = sum(1 for hr in res.homologated if hr.is_pseudocomponent)
        override_tag = " [H2O OVERRIDE]" if apply_override else ""
        print(f"  Esc={s.escenario}  C={str(s.corriente):<10}  "
              f"→ {res.n_used}/{n_compounds}  Σfrac={res.molar_sum_check:.6f}  "
              f"retenido={res.frac_retained*100:.1f}%  "
              f"pseudocomp={pseudo_count}  "
              f"total={res.total_mass_kgh:,.1f} kg/h"
              f"{override_tag}")

    # ── Post-homologation options (tóxicos) ───────────────────────────────
    flag_map = {c.phast_name: c for c in all_compounds}
    any_toxic_missing = any(
        any(flag_map.get(hr.phast_name) and flag_map[hr.phast_name].is_toxic
            for hr in res.homologated
            if hr.phast_name not in {cr.phast_name for cr in res.consolidated})
        for res in results
    )

    labeled_results: list[tuple[str, StreamResult]] = [("BASE", r) for r in results]

    if any_toxic_missing:
        print()
        print("⚠  Se detectaron compuestos tóxicos/ambos en el balance de entrada")
        print("   que no quedaron en el resultado homologado.")
        apply_opts = input("   ¿Aplicar opciones de ajuste post-homologación? [s/N]: ").strip().lower()
        if apply_opts in ('s', 'si', 'sí', 'y', 'yes'):
            labeled_results = apply_post_options(
                streams, results, all_compounds, hydrocarbon_only, aliases
            )

    print("\nGenerando Excel de salida...")
    write_output(labeled_results, output_path, n_compounds, all_compounds, water_threshold)
    print(f"\n✓ {len(results)} corrientes procesadas, N={n_compounds}")
    print(f"✓ Water Override aplicado a {len(override_indices)} corriente(s).")
    print(f"✓ Archivo guardado en: {output_path}")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    p = argparse.ArgumentParser(
        description='PHAST Homologator v2 — phast_pseudomps',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Uso básico (modo interactivo):
    python phast_homologator_v2.py

Uso silencioso:
    python phast_homologator_v2.py --input corrientes.xlsx --db phast_compounds.json
        --aliases aliases.json --output resultado.xlsx --n 14 --water-threshold 0.80

--water-threshold acepta valores decimales (0.80) o porcentajes (80).
Si se omite, se solicita interactivamente con default = 80%.
        """,
    )
    p.add_argument('--input',           default=None)
    p.add_argument('--db',              default=None)
    p.add_argument('--aliases',         default=None)
    p.add_argument('--output',          default=None)
    p.add_argument('--n',               type=int,   default=None)
    p.add_argument('--water-threshold', type=float, default=None,
                   dest='water_threshold',
                   help='Umbral fracción molar de agua para override (ej: 0.80 ó 80)')
    args = p.parse_args()

    # Normalise water threshold if given as percentage (e.g. 80 → 0.80)
    wt = args.water_threshold
    if wt is not None and wt > 1.0:
        wt /= 100.0

    SCRIPT_DIR = Path(__file__).parent

    print("=" * 60)
    print("  PHAST Homologator v2 — phast_pseudomps")
    print("=" * 60)
    print(f"\n  Directorio del script: {SCRIPT_DIR}")

    # 1. Input Excel
    if args.input:
        input_path = Path(args.input)
        if not input_path.is_absolute():
            input_path = SCRIPT_DIR / input_path
        if not input_path.exists():
            print(f"\n✗ Archivo de entrada no encontrado: {input_path}"); sys.exit(1)
    else:
        print("\n── Archivo de entrada (corrientes) ──────────────────────")
        input_path = prompt_file("EXCEL DE CORRIENTES", ('.xlsx',))

    # 2. PHAST DB JSON
    if args.db:
        db_path = Path(args.db)
        if not db_path.is_absolute():
            db_path = SCRIPT_DIR / args.db
        if not db_path.exists():
            print(f"\n✗ Base de datos no encontrada: {db_path}"); sys.exit(1)
    else:
        candidates = sorted(SCRIPT_DIR.glob('*.json'))
        phast_jsons = [f for f in candidates
                       if any(k in f.name.lower() for k in ('phast','compound','compounds'))]
        if len(phast_jsons) == 1:
            db_path = phast_jsons[0]
            print(f"\n── Base de datos PHAST ──────────────────────────────────")
            print(f"  Auto-detectado: {db_path.name}")
            raw = input("  [Enter para aceptar]: ").strip().strip('"').strip("'")
            if raw:
                db_path = Path(raw) if Path(raw).is_absolute() else SCRIPT_DIR / raw
        else:
            print("\n── Base de datos PHAST ──────────────────────────────────")
            db_path = prompt_file("BASE DE DATOS PHAST", ('.json',))

    # 3. Aliases JSON
    if args.aliases:
        aliases_path = Path(args.aliases)
        if not aliases_path.is_absolute():
            aliases_path = SCRIPT_DIR / args.aliases
        if not aliases_path.exists():
            print(f"\n✗ Archivo de alias no encontrado: {aliases_path}"); sys.exit(1)
    else:
        alias_jsons = [f for f in sorted(SCRIPT_DIR.glob('*.json'))
                       if any(k in f.name.lower() for k in ('alias','aliases'))]
        if len(alias_jsons) == 1:
            aliases_path = alias_jsons[0]
            print(f"\n── Tabla de alias ───────────────────────────────────────")
            print(f"  Auto-detectado: {aliases_path.name}")
            raw = input("  [Enter para aceptar]: ").strip().strip('"').strip("'")
            if raw:
                aliases_path = Path(raw) if Path(raw).is_absolute() else SCRIPT_DIR / raw
        else:
            print("\n── Tabla de alias ───────────────────────────────────────")
            aliases_path = prompt_file("TABLA DE ALIAS", ('.json',))

    # 4. Output path
    if args.output:
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = SCRIPT_DIR / args.output
    else:
        suggested = f"PHAST_homologation_v2_{input_path.stem}.xlsx"
        print(f"\n── Archivo de salida ────────────────────────────────────")
        output_path = prompt_output("resultado Excel", suggested)

    print()
    run(
        input_path      = input_path,
        db_path         = db_path,
        aliases_path    = aliases_path,
        output_path     = output_path,
        n_compounds     = args.n,
        water_threshold = wt,
    )
